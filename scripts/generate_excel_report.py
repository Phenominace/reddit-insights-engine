"""
Reddit Insights Engine - Excel Report Generator
Generates professional Excel reports from Reddit insight data
"""

import json
import subprocess
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

# Style definitions
TITLE_FONT = Font(name='Times New Roman', size=18, bold=True, color='000000')
HEADER_FONT = Font(name='Times New Roman', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1B3F66', end_color='1B3F66', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='3A5A8C', end_color='3A5A8C', fill_type='solid')
ALT_ROW_FILL = PatternFill(start_color='E9E9E9', end_color='E9E9E9', fill_type='solid')
BORDER = Border(
    left=Side(style='thin', color='E3DEDE'),
    right=Side(style='thin', color='E3DEDE'),
    top=Side(style='thin', color='E3DEDE'),
    bottom=Side(style='thin', color='E3DEDE')
)

def create_summary_sheet(wb: Workbook, report: dict) -> None:
    """Create the executive summary sheet."""
    ws = wb.active
    ws.title = "Executive Summary"
    
    # Title
    ws['B2'] = "Reddit Insights Report - Executive Summary"
    ws['B2'].font = TITLE_FONT
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 30
    
    # Report metadata
    row = 4
    ws[f'B{row}'] = "Report Generated:"
    ws[f'C{row}'] = datetime.fromisoformat(report['generatedAt'].replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M')
    ws[f'B{row+1}'] = "Date Range:"
    ws[f'C{row+1}'] = f"Last {report['configuration']['dateRange']['days']} days"
    ws[f'B{row+2}'] = "Subreddits Analyzed:"
    ws[f'C{row+2}'] = ", ".join(report['configuration']['subreddits'])
    
    # Key Metrics
    row = 9
    ws[f'B{row}'] = "KEY METRICS"
    ws[f'B{row}'].font = Font(name='Times New Roman', size=14, bold=True)
    ws.row_dimensions[row].height = 25
    
    metrics = [
        ("Total Posts Analyzed", report['summary']['totalPostsAnalyzed']),
        ("Total Pain Points Identified", report['summary']['totalPainPoints']),
        ("Total Questions Found", report['summary']['totalQuestions']),
        ("Content Opportunities", len(report['contentOpportunities'])),
    ]
    
    row += 1
    for label, value in metrics:
        ws[f'B{row}'] = label
        ws[f'C{row}'] = value
        ws[f'C{row}'].font = Font(name='Times New Roman', bold=True, color='0B5CAD')
        row += 1
    
    # Sentiment Distribution
    row += 1
    ws[f'B{row}'] = "SENTIMENT DISTRIBUTION"
    ws[f'B{row}'].font = Font(name='Times New Roman', size=14, bold=True)
    row += 1
    
    sentiment = report['summary']['sentimentDistribution']
    headers = ['Sentiment', 'Count', 'Percentage']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    total_sentiment = sum(sentiment.values())
    row += 1
    for sent, count in sentiment.items():
        ws.cell(row=row, column=2, value=sent.capitalize())
        ws.cell(row=row, column=3, value=count)
        ws.cell(row=row, column=4, value=f"{(count/total_sentiment*100):.1f}%" if total_sentiment > 0 else "0%")
        for col in range(2, 5):
            ws.cell(row=row, column=col).border = BORDER
        row += 1
    
    # Top Subreddits
    row += 1
    ws[f'B{row}'] = "TOP SUBREDDITS"
    ws[f'B{row}'].font = Font(name='Times New Roman', size=14, bold=True)
    row += 1
    
    headers = ['Subreddit', 'Posts Found']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    for sub in report['summary']['topSubreddits']:
        ws.cell(row=row, column=2, value=sub['name'])
        ws.cell(row=row, column=3, value=sub['count'])
        for col in range(2, 4):
            ws.cell(row=row, column=col).border = BORDER
        row += 1
    
    # Column widths
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15


def create_posts_sheet(wb: Workbook, report: dict) -> None:
    """Create the analyzed posts sheet."""
    ws = wb.create_sheet("Analyzed Posts")
    
    # Title
    ws['B2'] = "Analyzed Reddit Posts"
    ws['B2'].font = TITLE_FONT
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 30
    
    # Headers
    headers = ['Title', 'Subreddit', 'Sentiment', 'Opportunity', 'Summary', 'Pain Points', 'Questions', 'URL']
    row = 4
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Data
    row = 5
    for post in report['analyzedPosts']:
        ws.cell(row=row, column=2, value=post['title'][:100])
        ws.cell(row=row, column=3, value=post['subreddit'])
        ws.cell(row=row, column=4, value=post['sentiment'].capitalize())
        ws.cell(row=row, column=5, value=post['contentOpportunity'].capitalize())
        ws.cell(row=row, column=6, value=post['summary'][:150] if post.get('summary') else '')
        ws.cell(row=row, column=7, value=', '.join(post['painPoints'][:3]) if post.get('painPoints') else '')
        ws.cell(row=row, column=8, value=', '.join(post['questions'][:2]) if post.get('questions') else '')
        ws.cell(row=row, column=9, value=post['url'])
        
        # Apply alternating row fill and borders
        fill = ALT_ROW_FILL if row % 2 == 0 else None
        for col in range(2, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if fill:
                cell.fill = fill
        
        row += 1
    
    # Column widths
    widths = [50, 20, 12, 12, 40, 35, 35, 50]
    for col, width in enumerate(widths, start=2):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_opportunities_sheet(wb: Workbook, report: dict) -> None:
    """Create the content opportunities sheet."""
    ws = wb.create_sheet("Content Opportunities")
    
    # Title
    ws['B2'] = "Content Opportunities"
    ws['B2'].font = TITLE_FONT
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 30
    
    # Headers
    headers = ['Priority', 'Title', 'Description', 'Category', 'Content Types', 'Keywords']
    row = 4
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Data - sort by priority
    priority_order = {'high': 0, 'medium': 1, 'low': 2}
    opportunities = sorted(
        report['contentOpportunities'],
        key=lambda x: priority_order.get(x['priority'], 3)
    )
    
    row = 5
    for opp in opportunities:
        # Priority with color coding
        priority_cell = ws.cell(row=row, column=2, value=opp['priority'].upper())
        priority_cell.font = Font(name='Times New Roman', bold=True)
        if opp['priority'] == 'high':
            priority_cell.font = Font(name='Times New Roman', bold=True, color='FF0000')
        elif opp['priority'] == 'medium':
            priority_cell.font = Font(name='Times New Roman', bold=True, color='FFA500')
        
        ws.cell(row=row, column=3, value=opp['title'])
        ws.cell(row=row, column=4, value=opp['description'])
        ws.cell(row=row, column=5, value=opp['category'].replace('_', ' ').title())
        ws.cell(row=row, column=6, value=', '.join(opp['suggestedContentType']))
        ws.cell(row=row, column=7, value=', '.join(opp['keywords']))
        
        # Apply borders
        for col in range(2, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        row += 1
    
    # Column widths
    widths = [10, 40, 50, 15, 30, 35]
    for col, width in enumerate(widths, start=2):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_trending_sheet(wb: Workbook, report: dict) -> None:
    """Create the trending topics sheet."""
    ws = wb.create_sheet("Trending Topics")
    
    # Title
    ws['B2'] = "Trending Topics"
    ws['B2'].font = TITLE_FONT
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 30
    
    # Headers
    headers = ['Topic', 'Mentions', 'Dominant Sentiment']
    row = 4
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 5
    for topic in report['trendingTopics'][:30]:
        ws.cell(row=row, column=2, value=topic['topic'])
        ws.cell(row=row, column=3, value=topic['count'])
        ws.cell(row=row, column=4, value=topic['sentiment'].capitalize())
        
        for col in range(2, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            if row % 2 == 0:
                cell.fill = ALT_ROW_FILL
        
        row += 1
    
    # Column widths
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20


def create_audience_sheet(wb: Workbook, report: dict) -> None:
    """Create the audience insights sheet."""
    ws = wb.create_sheet("Audience Insights")
    
    # Title
    ws['B2'] = "Audience Personas & Pain Points"
    ws['B2'].font = TITLE_FONT
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 30
    
    # Headers
    headers = ['Audience Persona', 'Mentions', 'Top Pain Points']
    row = 4
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 5
    for audience in report['audienceInsights']:
        ws.cell(row=row, column=2, value=audience['persona'])
        ws.cell(row=row, column=3, value=audience['count'])
        ws.cell(row=row, column=4, value='\n'.join(f"• {pp}" for pp in audience['painPoints']))
        
        for col in range(2, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if row % 2 == 0:
                cell.fill = ALT_ROW_FILL
        
        ws.row_dimensions[row].height = 60
        row += 1
    
    # Column widths
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 60


def generate_excel_report(report_json: dict, output_path: str) -> str:
    """Generate Excel report from JSON data."""
    wb = Workbook()
    
    # Create all sheets
    create_summary_sheet(wb, report_json)
    create_posts_sheet(wb, report_json)
    create_opportunities_sheet(wb, report_json)
    create_trending_sheet(wb, report_json)
    create_audience_sheet(wb, report_json)
    
    # Save workbook
    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")
    
    return output_path


def main():
    """Main function to generate report from command line."""
    if len(sys.argv) < 2:
        print("Usage: python generate_excel_report.py <report_json_file> [output_path]")
        sys.exit(1)
    
    json_file = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else f"reddit_insights_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    with open(json_file, 'r') as f:
        report = json.load(f)
    
    generate_excel_report(report, output_path)


if __name__ == "__main__":
    main()
